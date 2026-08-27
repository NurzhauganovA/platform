"""Показ документа закупки прямо в платформе.

Раньше в браузере открывался только PDF, а `.docx` и `.xlsx` скачивались. Для
человека это выход из платформы: файл уезжает в загрузки, открывается чужой
программой, и обратно к строке он возвращается руками. За смену таких выходов
десятки — ТЗ смотрят по каждой закупке.

Документ не превращается в PDF: для этого нужен LibreOffice, а это полгигабайта
в образе ради предпросмотра. Вместо этого он разбирается на то, из чего
состоит, — абзацы и таблицы, — и рисуется браузером как обычная страница.
Точной вёрстки исходника при этом нет, и это честный размен: читают в ТЗ
требования и числа, а не поля страницы.

Наружу уходит разобранное, а не размётка. Собери здесь HTML — и содержимое
чужого документа стало бы кодом на нашей странице.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from platform_api.logging import get_logger

logger = get_logger(__name__)

MAX_BLOCKS = 600
"""Сколько абзацев и таблиц показывать. Техническое задание на тысячу страниц
браузер рисует минуту, а читают в нём первые страницы."""

MAX_ROWS = 300
MAX_COLUMNS = 40
"""Предел одного листа книги. В прайсе бывает тридцать тысяч строк, и целиком
он не нужен: смотрят состав и порядок величин."""

MAX_SHEETS = 12

_IMAGES = frozenset({".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp", ".tif", ".tiff"})
_PLAIN = frozenset({".txt", ".csv", ".md", ".log"})


@dataclass(frozen=True, slots=True)
class Block:
    """Кусок документа: заголовок, абзац или таблица."""

    kind: str
    """`heading`, `text` или `table`."""

    text: str = ""
    rows: tuple[tuple[str, ...], ...] = ()


@dataclass(frozen=True, slots=True)
class Sheet:
    """Лист книги."""

    title: str
    rows: tuple[tuple[str, ...], ...] = ()
    truncated: bool = False


@dataclass(frozen=True, slots=True)
class Preview:
    """Чем показывать документ."""

    kind: str
    """`pdf`, `image`, `document`, `sheet` или `none`."""

    blocks: tuple[Block, ...] = field(default_factory=tuple)
    sheets: tuple[Sheet, ...] = field(default_factory=tuple)
    truncated: bool = False
    note: str = ""
    """Почему показать нельзя. Молчание читается как поломка платформы."""


def build(path: Path) -> Preview:
    """Разбирает документ на то, чем его показать.

    Не падает ни на чём: битый файл, чужой формат, обрезанная выгрузка — всё
    это повод сказать словами и предложить скачать, а не отдать пятисотую.
    """
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return Preview(kind="pdf")
    if suffix in _IMAGES:
        return Preview(kind="image")

    try:
        if suffix == ".docx":
            return _document(path)
        if suffix in {".xlsx", ".xlsm"}:
            return _workbook(path)
        if suffix in _PLAIN:
            return _plain(path)
    except Exception as exc:
        logger.warning("Документ не разобрался", path=str(path), error=str(exc))
        return Preview(
            kind="none",
            note=(
                "Файл не удалось разобрать — возможно, он повреждён или "
                "сохранён в другом формате. Скачайте, чтобы открыть у себя."
            ),
        )

    return Preview(
        kind="none",
        note=(
            f"Формат «{suffix or 'без расширения'}» платформа показать не умеет. "
            "Старые «.doc» и «.xls» тоже: их читает только Word и Excel."
        ),
    )


def _document(path: Path) -> Preview:
    """Абзацы и таблицы документа Word — по порядку, как в файле.

    По порядку, а не сначала текст и потом таблицы: в техническом задании
    таблица идёт следом за своим заголовком, и разложенные порознь они теряют
    смысл.
    """
    from docx import Document
    from docx.document import Document as Doc
    from docx.oxml.ns import qn
    from docx.table import Table
    from docx.text.paragraph import Paragraph

    file: Doc = Document(str(path))
    blocks: list[Block] = []
    for child in file.element.body.iterchildren():
        if len(blocks) >= MAX_BLOCKS:
            return Preview(kind="document", blocks=tuple(blocks), truncated=True)
        if child.tag == qn("w:p"):
            paragraph = Paragraph(child, file)
            content = " ".join(paragraph.text.split())
            if content:
                style_name = (paragraph.style.name or "").lower() if paragraph.style else ""
                heading = "head" in style_name or "загол" in style_name
                blocks.append(Block(kind="heading" if heading else "text", text=content))
        elif child.tag == qn("w:tbl"):
            table = Table(child, file)
            lines = tuple(
                tuple(" ".join(cell.text.split()) for cell in row.cells[:MAX_COLUMNS])
                for row in table.rows[:MAX_ROWS]
            )
            if lines:
                blocks.append(Block(kind="table", rows=lines))
    return Preview(kind="document", blocks=tuple(blocks))


def _workbook(path: Path) -> Preview:
    """Листы книги Excel — значениями, а не формулами.

    Формула без пересчёта показала бы «=B2*C2» там, где человек ждёт сумму.
    """
    from openpyxl import load_workbook

    workbook = load_workbook(str(path), read_only=True, data_only=True)
    try:
        pages_out: list[Sheet] = []
        for label in workbook.sheetnames[:MAX_SHEETS]:
            sheet = workbook[label]
            lines: list[tuple[str, ...]] = []
            cut = False
            for number, row in enumerate(sheet.iter_rows(values_only=True)):
                if number >= MAX_ROWS:
                    cut = True
                    break
                lines.append(tuple(_cell(value) for value in row[:MAX_COLUMNS]))
            # Хвост пустых строк в книгах обычный: лист «на вырост». Из-за
            # него чтение упирается в предел, хотя содержимого три строки, —
            # и лист помечался обрезанным на ровном месте.
            while lines and not any(lines[-1]):
                lines.pop()
            pages_out.append(
                Sheet(
                    title=label,
                    rows=tuple(lines),
                    truncated=cut and len(lines) >= MAX_ROWS,
                )
            )
        return Preview(
            kind="sheet",
            sheets=tuple(pages_out),
            truncated=len(workbook.sheetnames) > MAX_SHEETS,
        )
    finally:
        workbook.close()


def _cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return " ".join(str(value).split())


def _plain(path: Path) -> Preview:
    """Простой текст — абзацами.

    Читается с заменой нечитаемых знаков, а не падает: выгрузки приходят и в
    windows-1251, и с обрывом посередине.
    """
    content = path.read_text(encoding="utf-8", errors="replace")
    paragraphs = [line.strip() for line in content.splitlines() if line.strip()]
    return Preview(
        kind="document",
        blocks=tuple(Block(kind="text", text=line) for line in paragraphs[:MAX_BLOCKS]),
        truncated=len(paragraphs) > MAX_BLOCKS,
    )


__all__ = ["MAX_BLOCKS", "MAX_ROWS", "Block", "Preview", "Sheet", "build"]
